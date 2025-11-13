import json
import csv
import io
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def generate_submission_pdf(submission):
    """Generate a PDF from a submission"""
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2196F3'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1976D2'),
        spaceAfter=8,
        spaceBefore=12
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6
    )
    
    # Title
    story.append(Paragraph("IITB SCAN - Form Submission", title_style))
    story.append(Spacer(1, 0.3*inch))
    
    # Submission metadata
    story.append(Paragraph("Submission Details", heading_style))
    metadata = [
        ['Field', 'Value'],
        ['Submission ID', str(submission.id)],
        ['Patient Name', submission.user.username],
        ['Patient Email', submission.user.email],
        ['Form Name', submission.form.name],
        ['Disease', submission.disease.name],
        ['Hospital', submission.hospital.name],
        ['Doctor', submission.doctor.name],
        ['Status', submission.status.title()],
        ['Submitted On', submission.created_at.strftime('%B %d, %Y at %H:%M:%S')],
        ['Updated On', submission.updated_at.strftime('%B %d, %Y at %H:%M:%S')],
    ]
    
    metadata_table = Table(metadata, colWidths=[2*inch, 4*inch])
    metadata_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E3F2FD')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    story.append(metadata_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Form data
    story.append(Paragraph("Form Responses", heading_style))
    submission_data = json.loads(submission.data)
    
    form_data = [['Field', 'Response']]
    for field in submission.form.fields:
        field_key = f'field_{field.field_name}'
        value = submission_data.get(field_key, 'Not provided')
        form_data.append([field.field_label, str(value)])
    
    form_table = Table(form_data, colWidths=[2.5*inch, 3.5*inch])
    form_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F5F5F5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(form_table)
    
    # Footer
    story.append(Spacer(1, 0.3*inch))
    footer_text = f"Generated on {datetime.now().strftime('%B %d, %Y at %H:%M:%S')} | IITB SCAN - Patient Data Collection System"
    story.append(Paragraph(footer_text, ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )))
    
    # Build PDF
    doc.build(story)
    pdf_buffer.seek(0)
    return pdf_buffer


def generate_submissions_csv(submissions, include_field_data=False):
    """Generate a CSV from multiple submissions"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header row
    if include_field_data:
        writer.writerow([
            'ID', 'Patient', 'Email', 'Disease', 'Hospital', 'Doctor', 'Form',
            'Status', 'Form Version', 'Created At', 'Updated At'
        ])
    else:
        writer.writerow([
            'ID', 'Patient', 'Email', 'Disease', 'Hospital', 'Doctor', 'Form',
            'Status', 'Created At', 'Updated At'
        ])
    
    # Data rows
    for submission in submissions:
        row = [
            submission.id,
            submission.user.username,
            submission.user.email,
            submission.disease.name,
            submission.hospital.name,
            submission.doctor.name,
            submission.form.name,
            submission.status,
            submission.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            submission.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        ]
        
        if not include_field_data:
            row.insert(8, submission.form_version)
        
        writer.writerow(row)
    
    output.seek(0)
    return output


def generate_detailed_submissions_csv(submissions):
    """Generate a detailed CSV with all form field responses in a clean, Excel-friendly format"""
    output = io.StringIO()
    writer = csv.writer(output)

    if not submissions:
        return output

    # Build a complete list of all possible fields across all submissions
    all_fields = {}
    for sub in submissions:
        for field in sub.form.fields:
            if field.field_label not in all_fields:
                all_fields[field.field_label] = {
                    'name': field.field_name,
                    'type': field.field_type,
                    'options': field.options
                }

    # Create header row with metadata and all possible fields
    header = [
        'Submission ID',
        'Patient Name',
        'Patient Email',
        'Disease',
        'Hospital',
        'Doctor',
        'Form Name',
        'Submission Status',
        'Submission Date',
        *[f'[{field_type.upper()}] {label}' for label, field in all_fields.items()]
    ]
    writer.writerow(header)

    # Write data rows
    for submission in submissions:
        try:
            # Handle both string and already-parsed JSON
            if isinstance(submission.data, str):
                submission_data = json.loads(submission.data)
            else:
                submission_data = submission.data
        except (json.JSONDecodeError, AttributeError) as e:
            print(f"Error parsing submission data: {e}")
            submission_data = {}
            
        # Start with metadata columns
        try:
            row = [
                submission.id,
                submission.user.username if hasattr(submission, 'user') else '',
                submission.user.email if hasattr(submission, 'user') else '',
                submission.disease.name if hasattr(submission, 'disease') and submission.disease else '',
                submission.hospital.name if hasattr(submission, 'hospital') and submission.hospital else '',
                submission.doctor.name if hasattr(submission, 'doctor') and submission.doctor else '',
                submission.form.name if hasattr(submission, 'form') and submission.form else '',
                getattr(submission, 'status', ''),
                submission.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(submission, 'created_at') else ''
            ]
        except Exception as e:
            print(f"Error getting submission metadata: {e}")
            row = [''] * 9  # Default empty row with 9 columns

        # Add field values in the same order as the header
        for label, field_info in all_fields.items():
            field_name = field_info['name']
            
            # Try different possible field key formats
            possible_keys = [
                f'field_{field_name}',  # field_name
                field_name,              # name (without field_ prefix)
                f'field_{field_name.lower()}',  # lowercase
                field_name.lower()              # lowercase without prefix
            ]
            
            value = ''
            for key in possible_keys:
                if key in submission_data:
                    value = submission_data[key]
                    break
            
            # Handle different field types
            if field_info['type'] in ['select', 'radio', 'checkbox'] and field_info['options']:
                try:
                    options = json.loads(field_info['options'])
                    if isinstance(options, list):
                        if field_info['type'] in ['select', 'radio'] and value:
                            # Find matching option for select/radio
                            for opt in options:
                                if str(opt.get('value', '')).strip().lower() == str(value).strip().lower():
                                    value = opt.get('label', value)
                                    break
                        elif field_info['type'] == 'checkbox' and value:
                            # Handle multiple selections for checkboxes
                            selected_values = [v.strip().lower() for v in str(value).split(',') if v.strip()]
                            selected_labels = []
                            for opt in options:
                                opt_value = str(opt.get('value', '')).strip().lower()
                                if opt_value in selected_values:
                                    selected_labels.append(opt.get('label', opt_value))
                            value = ', '.join(selected_labels) if selected_labels else value
                except (json.JSONDecodeError, AttributeError) as e:
                    print(f"Error processing options for field {field_name}: {e}")
            
            # Clean up the value for CSV
            if value is None:
                value = ''
            elif field_info['type'] == 'file':
                # For files, just indicate presence in CSV
                value = 'File Attached' if value else ''
            
            row.append(str(value).strip() if value is not None else '')

        writer.writerow(row)

    output.seek(0)
    return output


def generate_submissions_excel(submissions, include_field_data=False):
    """Generate an Excel file from submissions with improved formatting"""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")

    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.dimensions import ColumnDimension

    wb = Workbook()
    ws = wb.active
    ws.title = "Patient Submissions"

    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Alignment
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    if not include_field_data:
        # Basic export with just metadata
        headers = [
            'Submission ID', 'Patient Name', 'Email', 'Disease', 'Hospital', 
            'Doctor', 'Form Name', 'Status', 'Form Version', 'Submission Date'
        ]
        ws.append(headers)

        for submission in submissions:
            row = [
                submission.id,
                submission.user.username,
                submission.user.email,
                submission.disease.name,
                submission.hospital.name,
                submission.doctor.name,
                submission.form.name,
                submission.status,
                submission.form_version,
                submission.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ]
            ws.append(row)
    else:
        # Detailed export with form fields
        # Get all unique field labels across all forms
        all_fields = {}
        for sub in submissions:
            for field in sub.form.fields:
                if field.field_label not in all_fields:
                    all_fields[field.field_label] = {
                        'name': field.field_name,
                        'type': field.field_type,
                        'options': field.options
                    }

        # Create headers with field types
        headers = [
            'Submission ID',
            'Patient Name',
            'Email',
            'Disease',
            'Hospital',
            'Doctor',
            'Form Name',
            'Status',
            'Submission Date'
        ]
        
        # Add form field headers with their types
        for label, field in all_fields.items():
            headers.append(f"[{field['type'].upper()}] {label}")
        
        # Write headers
        ws.append(headers)

        # Add data rows
        for submission in submissions:
            try:
                submission_data = json.loads(submission.data)
            except (json.JSONDecodeError, AttributeError):
                submission_data = {}

            # Start with metadata columns
            row = [
                submission.id,
                submission.user.username,
                submission.user.email,
                submission.disease.name,
                submission.hospital.name,
                submission.doctor.name,
                submission.form.name,
                submission.status,
                submission.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ]

            # Add field values
            for label, field_info in all_fields.items():
                field_key = f'field_{field_info["name"]}'
                value = submission_data.get(field_key, '')
                
                # Handle different field types
                if field_info['type'] in ['select', 'radio', 'checkbox'] and field_info['options']:
                    try:
                        options = json.loads(field_info['options'])
                        if isinstance(options, list):
                            if field_info['type'] in ['select', 'radio'] and value:
                                # Find matching option for select/radio
                                for opt in options:
                                    if str(opt.get('value', '')).strip() == str(value).strip():
                                        value = opt.get('label', value)
                                        break
                            elif field_info['type'] == 'checkbox' and value:
                                # Handle multiple selections for checkboxes
                                selected_values = [v.strip() for v in str(value).split(',') if v.strip()]
                                selected_labels = []
                                for opt in options:
                                    opt_value = str(opt.get('value', '')).strip()
                                    if opt_value in selected_values:
                                        selected_labels.append(opt.get('label', opt_value))
                                value = ', '.join(selected_labels) if selected_labels else value
                    except (json.JSONDecodeError, AttributeError):
                        pass
                
                # Clean up the value
                if value is None:
                    value = ''
                row.append(str(value).strip())

            ws.append(row)

    # Apply styling to header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

    # Apply styling to data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = left_alignment

    # Auto-size columns with some reasonable limits
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        # Check header length first
        if column[0].row == 1:
            max_length = len(str(column[0].value or ''))
        
        # Check data cells
        for cell in column[1:]:  # Skip header
            try:
                cell_value = str(cell.value) if cell.value is not None else ''
                # Count newlines for wrapped text
                line_count = cell_value.count('\n') + 1
                max_line_length = max(len(line) for line in cell_value.split('\n'))
                
                # Consider both width and height for wrapped text
                cell_length = max(max_line_length, len(cell_value) // line_count)
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        
        # Set column width with limits
        adjusted_width = (max_length + 4) * 1.1  # Add some padding
        ws.column_dimensions[column_letter].width = min(max(adjusted_width, 10), 40)  # Min 10, Max 40
    
    ws.column_dimensions['J'].width = 18

    for col_num in range(11, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = 20

    # Freeze header row
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_detailed_submissions_excel(submissions):
    """Generate a detailed Excel file with all form field responses"""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")

    return generate_submissions_excel(submissions, include_field_data=True)
