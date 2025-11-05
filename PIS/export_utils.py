import json
import csv
import io
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_submission_pdf(submission):
    """Generate a PDF from a submission"""
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    story = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2196F3'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1976D2'),
        spaceAfter=8,
        spaceBefore=12
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6
    )
    
    # Title
    story.append(Paragraph("IITB SCAN - Form Submission", title_style))
    story.append(Spacer(1, 0.3*inch))
    
    # Submission metadata
    story.append(Paragraph("Submission Details", heading_style))
    metadata = [
        ['Field', 'Value'],
        ['Submission ID', str(submission.id)],
        ['Patient Name', submission.user.username],
        ['Patient Email', submission.user.email],
        ['Form Name', submission.form.name],
        ['Disease', submission.disease.name],
        ['Hospital', submission.hospital.name],
        ['Doctor', submission.doctor.name],
        ['Status', submission.status.title()],
        ['Submitted On', submission.created_at.strftime('%B %d, %Y at %H:%M:%S')],
        ['Updated On', submission.updated_at.strftime('%B %d, %Y at %H:%M:%S')],
    ]
    
    metadata_table = Table(metadata, colWidths=[2*inch, 4*inch])
    metadata_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E3F2FD')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    story.append(metadata_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Form data
    story.append(Paragraph("Form Responses", heading_style))
    submission_data = json.loads(submission.data)
    
    form_data = [['Field', 'Response']]
    for field in submission.form.fields:
        field_key = f'field_{field.field_name}'
        value = submission_data.get(field_key, 'Not provided')
        form_data.append([field.field_label, str(value)])
    
    form_table = Table(form_data, colWidths=[2.5*inch, 3.5*inch])
    form_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F5F5F5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    story.append(form_table)
    
    # Footer
    story.append(Spacer(1, 0.3*inch))
    footer_text = f"Generated on {datetime.now().strftime('%B %d, %Y at %H:%M:%S')} | IITB SCAN - Patient Data Collection System"
    story.append(Paragraph(footer_text, ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )))
    
    # Build PDF
    doc.build(story)
    pdf_buffer.seek(0)
    return pdf_buffer


def generate_submissions_csv(submissions, include_field_data=False):
    """Generate a CSV from multiple submissions"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header row
    if include_field_data:
        writer.writerow([
            'ID', 'Patient', 'Email', 'Form', 'Disease', 'Hospital', 'Doctor',
            'Status', 'Form Version', 'Created At', 'Updated At'
        ])
    else:
        writer.writerow([
            'ID', 'Patient', 'Email', 'Form', 'Disease', 'Hospital', 'Doctor',
            'Status', 'Created At', 'Updated At'
        ])
    
    # Data rows
    for submission in submissions:
        row = [
            submission.id,
            submission.user.username,
            submission.user.email,
            submission.form.name,
            submission.disease.name,
            submission.hospital.name,
            submission.doctor.name,
            submission.status,
            submission.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            submission.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        ]
        
        if not include_field_data:
            row.insert(8, submission.form_version)
        
        writer.writerow(row)
    
    output.seek(0)
    return output


def generate_detailed_submissions_csv(submissions):
    """Generate a detailed CSV with all form field responses"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    if not submissions:
        return output
    
    # Build header with all possible fields
    all_fields = set()
    for sub in submissions:
        for field in sub.form.fields:
            all_fields.add(field.field_label)
    
    header = ['ID', 'Patient', 'Email', 'Form', 'Disease', 'Hospital', 'Doctor', 'Status', 'Created At']
    header.extend(sorted(all_fields))
    writer.writerow(header)
    
    # Write data rows
    for submission in submissions:
        submission_data = json.loads(submission.data)
        row = [
            submission.id,
            submission.user.username,
            submission.user.email,
            submission.form.name,
            submission.disease.name,
            submission.hospital.name,
            submission.doctor.name,
            submission.status,
            submission.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ]
        
        # Add field data
        for field in submission.form.fields:
            field_key = f'field_{field.field_name}'
            value = submission_data.get(field_key, '')
            row.append(str(value))
        
        writer.writerow(row)
    
    output.seek(0)
    return output


def generate_submission_excel(submission):
    """Generate an Excel file from a single submission"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Submission {submission.id}"
    
    # Styling
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:B1')
    ws['A1'] = 'IITB SCAN - Form Submission'
    ws['A1'].font = Font(bold=True, size=14, color="2196F3")
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Submission Details Section
    ws['A3'] = 'Submission Details'
    ws['A3'].font = subheader_font
    ws['A3'].fill = subheader_fill
    ws.merge_cells('A3:B3')
    
    metadata = [
        ['Submission ID', str(submission.id)],
        ['Patient Name', submission.user.username],
        ['Patient Email', submission.user.email],
        ['Form Name', submission.form.name],
        ['Disease', submission.disease.name],
        ['Hospital', submission.hospital.name],
        ['Doctor', submission.doctor.name],
        ['Status', submission.status.title()],
        ['Submitted On', submission.created_at.strftime('%B %d, %Y at %H:%M:%S')],
        ['Updated On', submission.updated_at.strftime('%B %d, %Y at %H:%M:%S')],
    ]
    
    row = 4
    for field, value in metadata:
        ws[f'A{row}'] = field
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    # Form Responses Section
    row += 1
    ws[f'A{row}'] = 'Form Responses'
    ws[f'A{row}'].font = subheader_font
    ws[f'A{row}'].fill = subheader_fill
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    ws[f'A{row}'] = 'Field'
    ws[f'B{row}'] = 'Response'
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'A{row}'].border = border
    ws[f'B{row}'].border = border
    
    submission_data = json.loads(submission.data)
    row += 1
    for field in submission.form.fields:
        field_key = f'field_{field.field_name}'
        value = submission_data.get(field_key, 'Not provided')
        ws[f'A{row}'] = field.field_label
        ws[f'B{row}'] = str(value)
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    
    # Add footer
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = f"Generated on {datetime.now().strftime('%B %d, %Y at %H:%M:%S')} | IITB SCAN - Patient Data Collection System"
    ws[f'A{row}'].font = Font(size=9, color="808080", italic=True)
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


def generate_submissions_excel(submissions):
    """Generate an Excel file from multiple submissions with all form field data"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Submissions'
    
    if not submissions:
        return io.BytesIO()
    
    # Styling
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Build header with all possible fields
    all_fields = []
    field_set = set()
    for sub in submissions:
        for field in sub.form.fields:
            if field.field_label not in field_set:
                all_fields.append(field.field_label)
                field_set.add(field.field_label)
    
    header = ['ID', 'Patient', 'Email', 'Form', 'Disease', 'Hospital', 'Doctor', 'Status', 'Created At', 'Updated At']
    header.extend(all_fields)
    
    # Write header
    for col_num, header_text in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header_text
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data rows
    for row_num, submission in enumerate(submissions, 2):
        submission_data = json.loads(submission.data)
        
        # Basic info
        row_data = [
            submission.id,
            submission.user.username,
            submission.user.email,
            submission.form.name,
            submission.disease.name,
            submission.hospital.name,
            submission.doctor.name,
            submission.status.title(),
            submission.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            submission.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        ]
        
        # Add field data
        for field_label in all_fields:
            # Find the field in this submission's form
            value = ''
            for field in submission.form.fields:
                if field.field_label == field_label:
                    field_key = f'field_{field.field_name}'
                    value = submission_data.get(field_key, '')
                    break
            row_data.append(str(value))
        
        # Write row
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = border
            if col_num > 10:  # Form field columns
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    for col_num in range(1, len(header) + 1):
        column_letter = get_column_letter(col_num)
        if col_num <= 10:
            ws.column_dimensions[column_letter].width = 15
        else:
            ws.column_dimensions[column_letter].width = 20
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer
